import openpyxl

fileName='sales.xlsx'
wb=openpyxl.load_workbook(fileName,data_only=True)
#data only 防止讀取成公式


#print(wb.sheetnames)
#ws=wb['2020Q3']
ws=wb.worksheets[0]
# print(ws['E3'].value)
# print(ws.max_column)
# print(ws.max_row)

# for row in range(1,ws.max_row+1):
#     for  col in range(1,ws.max_column+1):
#         data=ws.cell(column=col,row=row).value
#         if data==None:
#             data=""
#         print("10%s"% data,end='')
#     print( )


# print(list(ws.columns)[1])
# for col in list(ws.columns):
#     for cell in col :
#         print(cell.value,end=' ')
#     print()

# for row in list(ws.rows):
#     for cell in row :
#         print(cell.value,end=' ')
#     print()

wb=openpyxl.Workbook()
ws=wb.active
wb.title="Super資料表"
wb.create_sheet(title='親組5K')
wb.create_sheet(title='X組5K')
# ws.create_sheet(title='T組5K')
wb.save('NFU.xlsx')